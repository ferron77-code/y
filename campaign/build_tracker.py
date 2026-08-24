"""Build the Cuba B459 podcast guest + influencer outreach tracker."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"

GOLD = "D9A441"
TEAL = "2E7D92"
BURGUNDY = "6B1D22"
INK = "141010"
CREAM = "F2EDE4"
YELLOW = "FFFF00"

hdr_fill = PatternFill("solid", fgColor=INK)
hdr_font = Font(name=FONT, size=10, bold=True, color=CREAM)
title_font = Font(name=FONT, size=16, bold=True, color=INK)
sub_font = Font(name=FONT, size=10, italic=True, color="666666")
body = Font(name=FONT, size=10)
bold = Font(name=FONT, size=10, bold=True)
input_fill = PatternFill("solid", fgColor=YELLOW)

thin = Side(style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ---------------------------------------------------------------- Read Me ---
ws = wb.active
ws.title = "Read Me"
ws.sheet_view.showGridLines = False

rows = [
    ("CUBA B459 — Podcast Guest & Influencer Outreach Tracker", "title"),
    ("Built for the 990 AM podcast launch. Work the Prospects tab; the Dashboard reads from it automatically.", "sub"),
    ("", None),
    ("HOW TO USE THIS FILE", "h"),
    ("1. Fill in one row per prospect on the Prospects tab. Yellow cells are the ones you type into.", "b"),
    ("2. Columns with a drop-down (Tier, Category, Language, Owner, Status, Channel) must use the list — the Dashboard counts on exact values.", "b"),
    ("3. Grey columns (Follow-Up Due, Days Since Contact) are formulas. Do not type in them.", "b"),
    ("4. Row 6 on Prospects is a filled-in EXAMPLE showing the expected format. Delete it once you start.", "b"),
    ("5. The Dashboard updates itself. Do not edit it.", "b"),
    ("", None),
    ("THE THREE TIERS", "h"),
    ("Tier 1 — Anchor", "b"),
    ("Nationally known voices. 100k+ reach. Expect a booking agent or press contact, and a low reply rate. Book 2–3 per quarter; each one legitimises the show for everyone below.", "n"),
    ("Tier 2 — Core", "b"),
    ("Journalists, musicians, academics, mid-size creators, 10k–100k. This is the engine of the show. Highest reply rate relative to reach. Aim for most bookings here.", "n"),
    ("Tier 3 — Community", "b"),
    ("Micro-creators, local Miami figures, exile-community organisers, under 10k. Reply and share at the highest rate, and their audiences are the most engaged. Cheapest reach you will ever get.", "n"),
    ("", None),
    ("WHERE TO FIND PUBLIC CONTACT DETAILS — legitimately", "h"),
    ("Do NOT scrape Instagram. It breaks Meta's terms, and mailing harvested addresses will get the CubaB459 domain flagged as spam right when you need it delivering.", "warn"),
    ("Use these instead. Every one is a contact the person has published because they WANT to be reached:", "b"),
    ("• The 'Email' button on an Instagram professional account — that address is published on purpose.", "n"),
    ("• Link-in-bio pages (Linktree, Beacons) — press kits and booking forms live there.", "n"),
    ("• The person's own website: /contact, /press, /booking, /epk.", "n"),
    ("• YouTube channel → About tab → 'View email address' (business enquiries).", "n"),
    ("• Substack / newsletter reply addresses.", "n"),
    ("• Newsroom staff directories for journalists.", "n"),
    ("• Booking agent or publicist named in a recent article or podcast appearance.", "n"),
    ("• Warm intros from Tony's radio network — 30 years and 24 stations. Work this FIRST; it beats every cold list.", "n"),
    ("", None),
    ("RULES THAT KEEP YOU OUT OF TROUBLE", "h"),
    ("• Every marketing email needs a real physical mailing address and a working unsubscribe link (CAN-SPAM). Use the station's address.", "n"),
    ("• A one-to-one, personally written booking invitation is not a marketing blast. Keep it that way: no BCC lists, no mail-merge blasts to cold addresses.", "n"),
    ("• Honour an opt-out within 10 business days, permanently.", "n"),
    ("• Paid social about Cuba policy is classed as social-issue advertising by Meta. It needs advertiser verification and a 'Paid for by' disclaimer BEFORE it will run. Start that verification early — it takes days.", "n"),
    ("• Sponsored content aired on 990 AM needs on-air sponsorship identification (FCC).", "n"),
    ("", None),
    ("CADENCE THAT WORKS", "h"),
    ("Touch 1 — the invitation. Touch 2 — follow up after 5 business days. Touch 3 — final note after another 7. Then stop and mark 'No Response'.", "n"),
    ("Three touches, then leave them alone. Someone who ignored three is not going to say yes to a fourth, and persistence past that costs you the relationship.", "n"),
]

r = 1
for text, kind in rows:
    c = ws.cell(row=r, column=1, value=text)
    if kind == "title":
        c.font = title_font
    elif kind == "sub":
        c.font = sub_font
    elif kind == "h":
        c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=TEAL)
    elif kind == "b":
        c.font = bold
    elif kind == "warn":
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BURGUNDY)
    else:
        c.font = body
    c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

ws.column_dimensions["A"].width = 118
for row in range(1, r):
    ws.row_dimensions[row].height = None

# -------------------------------------------------------------- Prospects ---
ps = wb.create_sheet("Prospects")
ps.sheet_view.showGridLines = False

COLS = [
    ("ID", 6, "auto"),
    ("Name", 24, "input"),
    ("Handle", 18, "input"),
    ("Platform", 13, "list"),
    ("Followers", 11, "input"),
    ("Tier", 15, "list"),
    ("Category", 16, "list"),
    ("Language", 12, "list"),
    ("Location", 16, "input"),
    ("Why Them (1 line)", 40, "input"),
    ("Public Contact", 26, "input"),
    ("Where Contact Came From", 22, "input"),
    ("Owner", 12, "list"),
    ("Status", 15, "list"),
    ("Date Contacted", 14, "input"),
    ("Date Replied", 13, "input"),
    ("Days Since Contact", 12, "calc"),
    ("Follow-Up Due", 13, "calc"),
    ("Channel", 11, "list"),
    ("Promoted Us?", 12, "list"),
    ("Notes", 44, "input"),
]

ps["A1"] = "PROSPECTS — one row per person"
ps["A1"].font = title_font
ps["A2"] = "Yellow = you type here.   Grey = formula, leave alone.   Row 6 is an example — delete it when you start."
ps["A2"].font = sub_font

HEAD_ROW = 4
for i, (name, width, _) in enumerate(COLS, start=1):
    c = ps.cell(row=HEAD_ROW, column=i, value=name)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = box
    ps.column_dimensions[get_column_letter(i)].width = width
ps.row_dimensions[HEAD_ROW].height = 30

FIRST, LAST = 5, 124

# Drop-down lists. Exact values — the Dashboard counts on them.
LISTS = {
    "Platform": ["Instagram", "X", "YouTube", "TikTok", "Podcast", "Facebook", "News Outlet"],
    "Tier": ["Tier 1 - Anchor", "Tier 2 - Core", "Tier 3 - Community"],
    "Category": ["Journalist", "Musician", "Activist", "Academic", "Comedian",
                 "Podcaster", "Faith Leader", "Business", "Politician", "Other"],
    "Language": ["Spanish", "English", "Bilingual"],
    "Owner": ["Tony", "Pedro", "Julio", "Ferron", "Unassigned"],
    "Status": ["Not Started", "Researching", "Contacted", "Replied", "Booked",
               "Recorded", "Published", "Declined", "No Response"],
    "Channel": ["Email", "DM", "Warm Intro", "Phone", "In Person"],
    "Promoted Us?": ["Yes", "No", "Not Yet"],
}

col_of = {name: i for i, (name, _, _) in enumerate(COLS, start=1)}

for name, options in LISTS.items():
    letter = get_column_letter(col_of[name])
    dv = DataValidation(
        type="list",
        formula1='"%s"' % ",".join(options),
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Pick a value from the list — the Dashboard counts exact values."
    dv.errorTitle = "Use the drop-down"
    ps.add_data_validation(dv)
    dv.add("%s%d:%s%d" % (letter, FIRST, letter, LAST))

grey = PatternFill("solid", fgColor="EDEDED")
input_cols = {i for i, (_, _, kind) in enumerate(COLS, start=1) if kind in ("input", "list")}
calc_cols = {i for i, (_, _, kind) in enumerate(COLS, start=1) if kind == "calc"}

dcol = get_column_letter(col_of["Date Contacted"])
rcol = get_column_letter(col_of["Date Replied"])

for row in range(FIRST, LAST + 1):
    # ID numbers the row so you can refer to a prospect in conversation.
    ps.cell(row=row, column=1, value="=IF(B%d=\"\",\"\",ROW()-%d)" % (row, FIRST - 1))

    # Days since contact — blank until contacted, stops counting once they reply.
    ps.cell(
        row=row,
        column=col_of["Days Since Contact"],
        value='=IF({d}{r}="","",IF({rp}{r}<>"","replied",TODAY()-{d}{r}))'.format(
            d=dcol, rp=rcol, r=row
        ),
    )
    # Follow-up due 5 business days after contact, per the cadence in Read Me.
    ps.cell(
        row=row,
        column=col_of["Follow-Up Due"],
        value='=IF({d}{r}="","",IF({rp}{r}<>"","-",{d}{r}+7))'.format(
            d=dcol, rp=rcol, r=row
        ),
    )

    for col in range(1, len(COLS) + 1):
        c = ps.cell(row=row, column=col)
        c.font = body
        c.border = box
        c.alignment = Alignment(vertical="top", wrap_text=(col in (10, 21)))
        if col in input_cols:
            c.fill = input_fill
        elif col in calc_cols or col == 1:
            c.fill = grey

for letter in (dcol, rcol, get_column_letter(col_of["Follow-Up Due"])):
    for row in range(FIRST, LAST + 1):
        ps["%s%d" % (letter, row)].number_format = "yyyy-mm-dd"
fcol = get_column_letter(col_of["Followers"])
for row in range(FIRST, LAST + 1):
    ps["%s%d" % (fcol, row)].number_format = "#,##0"

# One realistic example row so the expected format is unambiguous.
example = {
    "Name": "EXAMPLE — delete this row",
    "Handle": "@examplehandle",
    "Platform": "Instagram",
    "Followers": 48000,
    "Tier": "Tier 2 - Core",
    "Category": "Journalist",
    "Language": "Bilingual",
    "Location": "Miami, FL",
    "Why Them (1 line)": "Covers Cuban exile politics; interviewed 11J organisers in 2024.",
    "Public Contact": "press@example.com",
    "Where Contact Came From": "Press page on own site",
    "Owner": "Pedro",
    "Status": "Contacted",
    "Channel": "Email",
    "Promoted Us?": "Not Yet",
    "Notes": "Asked for a one-pager before committing. Send EPK + episode outline.",
}
from datetime import date, timedelta

ex_row = 6
for key, val in example.items():
    ps.cell(row=ex_row, column=col_of[key], value=val)
ps.cell(row=ex_row, column=col_of["Date Contacted"], value=date.today() - timedelta(days=3))
ps["%s%d" % (dcol, ex_row)].number_format = "yyyy-mm-dd"
for col in range(1, len(COLS) + 1):
    ps.cell(row=ex_row, column=col).font = Font(name=FONT, size=10, italic=True, color="808080")

ps.freeze_panes = "C5"
ps.auto_filter.ref = "A%d:%s%d" % (HEAD_ROW, get_column_letter(len(COLS)), LAST)

# -------------------------------------------------------------- Dashboard ---
db = wb.create_sheet("Dashboard")
db.sheet_view.showGridLines = False
db["A1"] = "PIPELINE DASHBOARD"
db["A1"].font = title_font
db["A2"] = "Reads from Prospects automatically. Do not type on this sheet."
db["A2"].font = sub_font

scol = get_column_letter(col_of["Status"])
tcol = get_column_letter(col_of["Tier"])
pcol = get_column_letter(col_of["Promoted Us?"])
srange = "Prospects!$%s$%d:$%s$%d" % (scol, FIRST, scol, LAST)
trange = "Prospects!$%s$%d:$%s$%d" % (tcol, FIRST, tcol, LAST)
prange = "Prospects!$%s$%d:$%s$%d" % (pcol, FIRST, pcol, LAST)
nrange = "Prospects!$B$%d:$B$%d" % (FIRST, LAST)


def block(start_row, title, pairs):
    c = db.cell(row=start_row, column=1, value=title)
    c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=TEAL)
    db.cell(row=start_row, column=2).fill = PatternFill("solid", fgColor=TEAL)
    r = start_row + 1
    for label, formula in pairs:
        lc = db.cell(row=r, column=1, value=label)
        lc.font = body
        lc.border = box
        vc = db.cell(row=r, column=2, value=formula)
        vc.font = bold
        vc.border = box
        vc.alignment = Alignment(horizontal="center")
        r += 1
    return r + 1


row = 4
row = block(row, "PIPELINE BY STATUS", [
    ("Total prospects entered", '=COUNTIF(%s,"<>")' % nrange),
] + [
    (s, '=COUNTIF(%s,"%s")' % (srange, s)) for s in LISTS["Status"]
])

row = block(row, "BY TIER", [
    (t, '=COUNTIF(%s,"%s")' % (trange, t)) for t in LISTS["Tier"]
])

reached = '(COUNTIF({s},"Contacted")+COUNTIF({s},"Replied")+COUNTIF({s},"Booked")+COUNTIF({s},"Recorded")+COUNTIF({s},"Published")+COUNTIF({s},"Declined")+COUNTIF({s},"No Response"))'.format(s=srange)
replied = '(COUNTIF({s},"Replied")+COUNTIF({s},"Booked")+COUNTIF({s},"Recorded")+COUNTIF({s},"Published")+COUNTIF({s},"Declined"))'.format(s=srange)
booked = '(COUNTIF({s},"Booked")+COUNTIF({s},"Recorded")+COUNTIF({s},"Published"))'.format(s=srange)

row = block(row, "CONVERSION", [
    ("Outreach sent", "=%s" % reached),
    ("Replies received", "=%s" % replied),
    ("Reply rate", "=IFERROR(%s/%s,0)" % (replied, reached)),
    ("Booked or better", "=%s" % booked),
    ("Book rate (of those contacted)", "=IFERROR(%s/%s,0)" % (booked, reached)),
    ("Episodes published", '=COUNTIF(%s,"Published")' % srange),
    ("Guests who promoted us", '=COUNTIF(%s,"Yes")' % prange),
])

row = block(row, "NEEDS ACTION NOW", [
    ("Follow-ups due today or overdue",
     '=COUNTIFS(Prospects!${fu}${a}:${fu}${b},">0",Prospects!${fu}${a}:${fu}${b},"<="&TODAY())'.format(
         fu=get_column_letter(col_of["Follow-Up Due"]), a=FIRST, b=LAST)),
    ("Researched but never contacted", '=COUNTIF(%s,"Researching")' % srange),
    ("Unassigned owner", '=COUNTIF(Prospects!$%s$%d:$%s$%d,"Unassigned")' % (
        get_column_letter(col_of["Owner"]), FIRST, get_column_letter(col_of["Owner"]), LAST)),
])

db.column_dimensions["A"].width = 36
db.column_dimensions["B"].width = 14

# Percentages as fractions.
for rr in range(1, row + 2):
    label = db.cell(row=rr, column=1).value
    if isinstance(label, str) and "rate" in label.lower():
        db.cell(row=rr, column=2).number_format = "0.0%"

wb.save("/home/user/y/campaign/CubaB459-Outreach-Tracker.xlsx")
print("written")
